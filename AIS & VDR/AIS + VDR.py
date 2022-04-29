import openpyxl
from time import sleep

# Excel sheet path specification
AISdata = openpyxl.load_workbook("AISdata.xlsx")
sh = AISdata.active

# Loop thru data sheet
with open("voyageData.txt", 'a+') as VDR:
    for i in range(2, sh.max_row + 1):
        for j in range(1, sh.max_column + 1):
            cell_obj = sh.cell(row=i, column=j)
            AISdataString = str(cell_obj.value) + "\n"
            print(AISdataString)
            VDR.write(AISdataString)
        sleep(6)
VDR.close()